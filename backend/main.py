import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from time import time
from typing import Optional

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(__file__))))
from env_loader import load_env
load_env()

from fastapi import FastAPI, HTTPException, Depends, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.hash import bcrypt
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

from database import get_conn, init_db, seed_admin, new_token

# ── Config ──────────────────────────────────────────────────────────────────
SECRET_KEY = os.getenv("SECRET_KEY", "")
ALGORITHM  = "HS256"
TOKEN_EXP_HOURS = 8

if not SECRET_KEY:
    raise RuntimeError("SECRET_KEY belum di-set di .env")

VALID_UNITS = ("MI", "MTs", "RA", "ALL")
VALID_STATUS = ("Hadir", "Izin", "Sakit")

LOGIN_MAX_ATTEMPTS = 5
LOGIN_MAX_ATTEMPTS_IP = 10
LOGIN_WINDOW_SEC = 60

# ── Excel formula injection guard ────────────────────────────────────────────
_FORMULA_PREFIX = ("=", "+", "-", "@", "\t", "\r")

def excel_safe(value):
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIX):
        return "'" + value
    return value

# ── App ──────────────────────────────────────────────────────────────────────
app = FastAPI(title="Absensi Murid API", docs_url=None, redoc_url=None, openapi_url=None)

CORS_ORIGINS = os.getenv(
    "CORS_ORIGINS", "http://localhost:8000,http://127.0.0.1:8000"
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in CORS_ORIGINS.split(",") if o.strip()],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def force_https(request: Request, call_next):
    proto = request.headers.get("x-forwarded-proto")
    if proto and proto.split(",")[0].strip().lower() != "https":
        return RedirectResponse(request.url.replace(scheme="https"), status_code=301)
    return await call_next(request)

@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "no-referrer"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "img-src 'self' data: https://maps.googleapis.com https://*.googleapis.com https://*.gstatic.com; "
        "style-src 'self' 'unsafe-inline'; "
        "script-src 'self'; "
        "connect-src 'self' https://*.googleapis.com; "
        "font-src 'self' data:; "
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self'"
    )
    response.headers["Permissions-Policy"] = (
        "camera=(), microphone=(), geolocation=(), "
        "payment=(), usb=(), battery=(), "
        "fullscreen=(self), interest-cohort=()"
    )
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    response.headers["Cross-Origin-Opener-Policy"] = "same-origin"
    response.headers["Cross-Origin-Resource-Policy"] = "same-origin"
    response.headers["Cross-Origin-Embedder-Policy"] = "require-corp"
    return response

# ── JWT Helpers ──────────────────────────────────────────────────────────────
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/login")

def create_token(data: dict) -> str:
    payload = data.copy()
    payload["exp"] = datetime.utcnow() + timedelta(hours=TOKEN_EXP_HOURS)
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

def verify_token(token: str = Depends(oauth2_scheme)) -> dict:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError:
        raise HTTPException(status_code=401, detail="Token tidak valid atau expired")

    username = payload.get("sub")
    unit     = payload.get("unit")
    if not username or not unit:
        raise HTTPException(status_code=401, detail="Token tidak valid atau expired")

    # Validasi unit terhadap database agar admin tidak bisa escalate unit
    # (mis. admin_mi memalsukan unit=ALL / MTs lewat token).
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute("SELECT unit FROM admin WHERE username = ?", (username,))
    row  = cur.fetchone()
    conn.close()

    if not row:
        raise HTTPException(status_code=401, detail="Token tidak valid atau expired")

    db_unit = row["unit"]
    if unit != db_unit and db_unit != "ALL":
        raise HTTPException(status_code=403, detail="Unit tidak sesuai dengan akun")

    payload["unit"] = db_unit
    return payload

_login_attempts = defaultdict(list)
_login_ip_attempts = defaultdict(list)

def check_login_attempts(request: Request, username: str):
    now = time()
    ip  = request.client.host

    if len(_login_attempts) > 2000:
        cutoff = now - LOGIN_WINDOW_SEC
        for d in (_login_attempts, _login_ip_attempts):
            for k in [k for k, v in d.items() if not v or v[-1] < cutoff]:
                del d[k]

    key = f"{ip}:{username}"
    _login_attempts[key] = [t for t in _login_attempts[key] if now - t < LOGIN_WINDOW_SEC]
    if len(_login_attempts[key]) >= LOGIN_MAX_ATTEMPTS:
        raise HTTPException(status_code=429, detail="Terlalu banyak percobaan login. Coba lagi nanti.")

    _login_ip_attempts[ip] = [t for t in _login_ip_attempts[ip] if now - t < LOGIN_WINDOW_SEC]
    if len(_login_ip_attempts[ip]) >= LOGIN_MAX_ATTEMPTS_IP:
        raise HTTPException(status_code=429, detail="Terlalu banyak percobaan login. Coba lagi nanti.")

    return ip, key

# ── Rate limit absen QR (per token + IP) ──────────────────────────────────────
ABSEB_MAX_PER_MIN = 5
_absen_attempts = defaultdict(list)

def check_absen_rate_limit(request: Request, token: str):
    now  = time()
    ip   = request.client.host
    key  = f"{ip}:{token}"
    _absen_attempts[key] = [t for t in _absen_attempts[key] if now - t < 60]
    if len(_absen_attempts[key]) >= ABSEB_MAX_PER_MIN:
        raise HTTPException(status_code=429, detail="Terlalu banyak percobaan absen. Tunggu sebentar.")
    _absen_attempts[key].append(now)
    if len(_absen_attempts) > 5000:
        cutoff = now - 60
        for k in [k for k, v in _absen_attempts.items() if not v or v[-1] < cutoff]:
            del _absen_attempts[k]

# ── Startup ──────────────────────────────────────────────────────────────────
@app.on_event("startup")
def startup():
    init_db()
    seed_admin()

# ── Auth ─────────────────────────────────────────────────────────────────────
@app.post("/api/login")
def login(request: Request, form: OAuth2PasswordRequestForm = Depends()):
    ip, key = check_login_attempts(request, form.username)

    conn = get_conn()
    cur  = conn.cursor()
    cur.execute("SELECT * FROM admin WHERE username = ?", (form.username,))
    admin = cur.fetchone()
    conn.close()

    # ponytail: dummy hash biar waktu verifikasi seragam — user yang tidak ada
    # tetap jalan bcrypt sehingga attacker tidak bisa bedakan via timing.
    DUMMY = "$2b$12$vt5g/54OZcsrT9MsRc4W2u88bwcOYvHv6FNb1Rx/f88r9Z5zzPDAW"
    pwhash = admin["password_hash"] if admin else DUMMY
    if not admin or not bcrypt.verify(form.password, pwhash):
        _login_attempts[key].append(time())
        _login_ip_attempts[ip].append(time())
        raise HTTPException(status_code=401, detail="Username atau password salah")

    _login_attempts.pop(key, None)
    _login_ip_attempts.pop(ip, None)
    token = create_token({"sub": admin["username"], "unit": admin["unit"]})
    return {
        "access_token": token,
        "token_type": "bearer",
        "unit": admin["unit"],
        "username": admin["username"],
    }

@app.get("/api/me")
def me(payload: dict = Depends(verify_token)):
    return {"username": payload["sub"], "unit": payload["unit"]}

# ── Admin: kelola murid + QR ────────────────────────────────────────────────
class MuridIn(BaseModel):
    nama: str
    kelas: str
    unit: str
    urutan: int = 0

@app.get("/api/murid")
def list_murid(payload: dict = Depends(verify_token)):
    conn = get_conn()
    cur  = conn.cursor()
    if payload["unit"] == "ALL":
        cur.execute("SELECT id, nama, kelas, unit, urutan, token FROM murid ORDER BY urutan, nama")
    else:
        cur.execute("SELECT id, nama, kelas, unit, urutan, token FROM murid WHERE unit = ? ORDER BY urutan, nama",
                    (payload["unit"],))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

@app.post("/api/murid")
def create_murid(data: MuridIn, payload: dict = Depends(verify_token)):
    data.unit = payload["unit"] if payload["unit"] != "ALL" else data.unit
    if data.unit not in VALID_UNITS or data.unit == "ALL":
        raise HTTPException(status_code=400, detail="Unit tidak valid")
    conn = get_conn()
    cur  = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO murid (nama, kelas, unit, urutan, token) VALUES (?, ?, ?, ?, ?)",
            (data.nama, data.kelas, data.unit, data.urutan, new_token()))
        conn.commit()
        mid = cur.lastrowid
    except Exception:
        raise HTTPException(status_code=400, detail="Data tidak valid")
    finally:
        conn.close()
    return {"ok": True, "id": mid, "token": None}

@app.delete("/api/murid/{murid_id}")
def delete_murid(murid_id: int, payload: dict = Depends(verify_token)):
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute(
        "DELETE FROM murid WHERE id = ? AND (unit = ? OR ? = 'ALL')",
        (murid_id, payload["unit"], payload["unit"]),
    )
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    if not deleted:
        raise HTTPException(status_code=404, detail="Murid tidak ditemukan")
    return {"ok": True, "deleted": deleted}

@app.post("/api/murid/{murid_id}/reset-token")
def reset_murid_token(murid_id: int, payload: dict = Depends(verify_token)):
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute(
        "UPDATE murid SET token = ? WHERE id = ? AND (unit = ? OR ? = 'ALL')",
        (new_token(), murid_id, payload["unit"], payload["unit"]),
    )
    conn.commit()
    updated = cur.rowcount
    conn.close()
    if not updated:
        raise HTTPException(status_code=404, detail="Murid tidak ditemukan")
    return {"ok": True}

@app.get("/api/murid/{murid_id}/qr")
def murid_qr(murid_id: int, payload: dict = Depends(verify_token)):
    import qrcode
    from qrcode.constants import ERROR_CORRECT_H
    from PIL import Image
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute(
        "SELECT token FROM murid WHERE id = ? AND (unit = ? OR ? = 'ALL')",
        (murid_id, payload["unit"], payload["unit"]),
    )
    row = cur.fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Murid tidak ditemukan")
    url = f"{os.getenv('PUBLIC_URL', 'http://localhost:8000')}/absen?t={row['token']}"
    qr = qrcode.QRCode(error_correction=ERROR_CORRECT_H, box_size=10, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    # Logo di atas latar putih solid. kotak ~1/3 lebar QR.
    # ponytail: error-correction H cuma tahan ~30% area tertutup; gedein lagi = gagal scan.
    logo_path = os.path.join(os.path.dirname(__file__), "dashboard", "static", "img", "logo.png")
    if os.path.exists(logo_path):
        logo = Image.open(logo_path).convert("RGBA")
        pad = img.size[0] // 3
        logo.thumbnail((int(pad * 0.8), int(pad * 0.8)))

        box = Image.new("RGBA", (pad, pad), (255, 255, 255, 255))
        pos = ((img.size[0] - pad) // 2, (img.size[1] - pad) // 2)
        box.paste(logo, ((pad - logo.width) // 2, (pad - logo.height) // 2), logo)
        img.paste(box, pos)

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/png")

# ── Halaman absen publik: lookup token → identitas murid ───────────────────
@app.get("/api/absen/info/{token}")
def absen_info(token: str):
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute("SELECT id, nama, kelas, unit, urutan FROM murid WHERE token = ?", (token,))
    row = cur.fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="QR tidak valid")
    return dict(row)

# ── Absen QR (web form) ─────────────────────────────────────────────────────
@app.post("/api/absen-web")
async def absen_web(
    request: Request,
    token: str    = Form(...),
    status:  str  = Form(...),
    latitude:  Optional[float] = Form(None),
    longitude: Optional[float] = Form(None),
):
    # ponytail: QR per-murid = token statis tanpa waktu kedaluwarsa.
    # Ditutup: rate-limit per token+IP, dan maksimal 1 absen/hari per murid
    # (index unik idx_absen_1perday). Fotonya QR masih bisa di-titipin sampai
    # habis kuota harian — upgrade nyata: token sekali pakai / berjangka.
    if status not in VALID_STATUS:
        raise HTTPException(status_code=400, detail="Status absensi tidak valid")

    check_absen_rate_limit(request, token)

    conn = get_conn()
    cur  = conn.cursor()
    cur.execute("SELECT * FROM murid WHERE token = ?", (token,))
    murid = cur.fetchone()
    if not murid:
        conn.close()
        raise HTTPException(status_code=404, detail="QR tidak valid")

    now    = datetime.now()
    jam    = now.strftime("%H:%M")
    tanggal = now.strftime("%Y-%m-%d")

    cur.execute("SELECT id FROM absensi WHERE murid_id = ? AND tanggal = ?", (murid["id"], tanggal))
    if cur.fetchone():
        conn.close()
        raise HTTPException(status_code=409, detail="Absen hari ini sudah tercatat")

    cur.execute("""
        INSERT INTO absensi (murid_id, nama, kelas, unit, urutan, status, latitude, longitude, jam, tanggal)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (murid["id"], murid["nama"], murid["kelas"], murid["unit"], murid["urutan"],
          status, latitude, longitude, jam, tanggal))
    conn.commit()
    absensi_id = cur.lastrowid
    conn.close()

    return {
        "ok": True,
        "id": absensi_id,
        "nama": murid["nama"],
        "kelas": murid["kelas"],
        "unit": murid["unit"],
        "status": status,
        "jam": jam,
        "tanggal": tanggal,
    }

# ── Dashboard - Baca Absensi ─────────────────────────────────────────────────
@app.get("/api/absensi")
def list_absensi(
    tanggal: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    unit = payload["unit"]
    if unit not in VALID_UNITS:
        raise HTTPException(status_code=403, detail="Unit tidak valid")
    conn = get_conn()
    cur  = conn.cursor()

    if tanggal:
        if unit == "ALL":
            cur.execute("SELECT * FROM absensi WHERE tanggal = ? ORDER BY urutan, id", (tanggal,))
        else:
            cur.execute("SELECT * FROM absensi WHERE unit = ? AND tanggal = ? ORDER BY urutan, id",
                        (unit, tanggal))
    else:
        if unit == "ALL":
            cur.execute("SELECT * FROM absensi ORDER BY tanggal DESC, urutan, id LIMIT 200")
        else:
            cur.execute("SELECT * FROM absensi WHERE unit = ? ORDER BY tanggal DESC, urutan, id LIMIT 200", (unit,))

    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

# ── Dashboard - Hapus Absensi ─────────────────────────────────────────────────
class DeleteIds(BaseModel):
    ids: list[int]

@app.delete("/api/absensi")
def delete_absensi(data: DeleteIds, payload: dict = Depends(verify_token)):
    if not data.ids:
        raise HTTPException(status_code=400, detail="Tidak ada ID yang dikirim")

    conn = get_conn()
    cur  = conn.cursor()

    placeholders = ",".join("?" * len(data.ids))
    cur.execute(
        f"DELETE FROM absensi WHERE id IN ({placeholders}) AND (unit = ? OR ? = 'ALL')",
        (*data.ids, payload["unit"], payload["unit"]),
    )
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    return {"ok": True, "deleted": deleted}

# ── Export Excel ──────────────────────────────────────────────────────────────
@app.get("/api/absensi/export")
def export_excel(
    tanggal: Optional[str] = None,
    payload: dict = Depends(verify_token)
):
    unit = payload["unit"]
    if unit not in VALID_UNITS:
        raise HTTPException(status_code=403, detail="Unit tidak valid")
    conn = get_conn()
    cur  = conn.cursor()

    if tanggal:
        if unit == "ALL":
            cur.execute("SELECT * FROM absensi WHERE tanggal = ? ORDER BY urutan, id", (tanggal,))
        else:
            cur.execute("SELECT * FROM absensi WHERE unit = ? AND tanggal = ? ORDER BY urutan, id",
                        (unit, tanggal))
    else:
        if unit == "ALL":
            cur.execute("SELECT * FROM absensi ORDER BY tanggal DESC, urutan, id")
        else:
            cur.execute("SELECT * FROM absensi WHERE unit = ? ORDER BY tanggal DESC, urutan, id", (unit,))

    rows = cur.fetchall()
    conn.close()

    # Buat Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    unit_label = unit if unit != "ALL" else "Semua Unit"
    ws.title   = f"Absensi {unit_label}"

    # Style
    header_fill   = PatternFill("solid", fgColor="1E3A5F")
    header_font   = Font(color="FFFFFF", bold=True, size=11)
    center_align  = Alignment(horizontal="center", vertical="center")
    border_side   = Side(style="thin", color="CCCCCC")
    cell_border   = Border(left=border_side, right=border_side,
                           top=border_side, bottom=border_side)

    # Judul
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value     = f"Rekap Absensi Murid {unit_label}"
    title_cell.font      = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment = center_align
    if tanggal:
        ws.merge_cells("A2:I2")
        ws["A2"].value     = f"Tanggal: {tanggal}"
        ws["A2"].alignment = center_align
        ws["A2"].font      = Font(italic=True, color="666666")

    # Header tabel
    headers = ["No", "Tanggal", "Jam", "Nama Murid", "Kelas",
               "Unit", "Status", "Latitude", "Longitude"]
    start_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = cell_border

    # Data
    status_colors = {"Hadir": "D4EDDA", "Izin": "FFF3CD", "Sakit": "F8D7DA"}
    for i, row in enumerate(rows, 1):
        r = start_row + i
        values = [
            i, excel_safe(row["tanggal"]), excel_safe(row["jam"]),
            excel_safe(row["nama"]), excel_safe(row["kelas"]),
            excel_safe(row["unit"]), excel_safe(row["status"]),
            row["latitude"] or "-", row["longitude"] or "-"
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = center_align
            cell.border    = cell_border
            if col == 7:  # status
                color = status_colors.get(row["status"], "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)

    # Lebar kolom
    col_widths = [5, 12, 8, 25, 22, 8, 10, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = f"A{start_row + 1}"

    # Stream ke response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"absensi_{unit_label}_{tanggal or 'semua'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )

# ── Stats ─────────────────────────────────────────────────────────────────────
@app.get("/api/stats")
def stats(tanggal: Optional[str] = None, payload: dict = Depends(verify_token)):
    unit = payload["unit"]
    if unit not in VALID_UNITS:
        raise HTTPException(status_code=403, detail="Unit tidak valid")
    conn = get_conn()
    cur  = conn.cursor()

    params = []
    where  = []
    if unit != "ALL":
        where.append("unit = ?")
        params.append(unit)
    if tanggal:
        where.append("tanggal = ?")
        params.append(tanggal)
    where_sql = " AND ".join(where) if where else "1=1"

    cur.execute(f"""
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status='Hadir' THEN 1 ELSE 0 END) as hadir,
            SUM(CASE WHEN status='Izin'  THEN 1 ELSE 0 END) as izin,
            SUM(CASE WHEN status='Sakit' THEN 1 ELSE 0 END) as sakit
        FROM absensi WHERE {where_sql}
    """, params)
    row = dict(cur.fetchone())
    conn.close()
    return row

# ── Rekap per Murid ──────────────────────────────────────────────────────────
@app.get("/api/rekap")
def rekap(
    start: Optional[str] = None,
    end: Optional[str] = None,
    payload: dict = Depends(verify_token),
):
    unit = payload["unit"]
    if unit not in VALID_UNITS:
        raise HTTPException(status_code=403, detail="Unit tidak valid")
    conn = get_conn()
    cur  = conn.cursor()

    params = []
    where  = []
    if unit != "ALL":
        where.append("unit = ?")
        params.append(unit)
    if start:
        where.append("tanggal >= ?")
        params.append(start)
    if end:
        where.append("tanggal <= ?")
        params.append(end)
    where_sql = " AND ".join(where) if where else "1=1"

    cur.execute(f"""
        SELECT nama, kelas, unit, MIN(urutan) as urutan,
               COUNT(*) as total,
               SUM(CASE WHEN status='Hadir' THEN 1 ELSE 0 END) as hadir,
               SUM(CASE WHEN status='Izin'  THEN 1 ELSE 0 END) as izin,
               SUM(CASE WHEN status='Sakit' THEN 1 ELSE 0 END) as sakit
        FROM absensi WHERE {where_sql}
        GROUP BY nama, kelas, unit
        ORDER BY urutan, nama ASC
    """, params)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

@app.get("/api/rekap/export")
def export_rekap_excel(
    start: Optional[str] = None,
    end: Optional[str] = None,
    payload: dict = Depends(verify_token),
):
    unit = payload["unit"]
    if unit not in VALID_UNITS:
        raise HTTPException(status_code=403, detail="Unit tidak valid")
    conn = get_conn()
    cur  = conn.cursor()

    params = []
    where  = []
    if unit != "ALL":
        where.append("unit = ?")
        params.append(unit)
    if start:
        where.append("tanggal >= ?")
        params.append(start)
    if end:
        where.append("tanggal <= ?")
        params.append(end)
    where_sql = " AND ".join(where) if where else "1=1"

    cur.execute(f"""
        SELECT nama, kelas, unit, MIN(urutan) as urutan,
               COUNT(*) as total,
               SUM(CASE WHEN status='Hadir' THEN 1 ELSE 0 END) as hadir,
               SUM(CASE WHEN status='Izin'  THEN 1 ELSE 0 END) as izin,
               SUM(CASE WHEN status='Sakit' THEN 1 ELSE 0 END) as sakit
        FROM absensi WHERE {where_sql}
        GROUP BY nama, kelas, unit
        ORDER BY urutan, nama ASC
    """, params)
    rows = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active

    unit_label = unit if unit != "ALL" else "Semua Unit"
    ws.title   = f"Rekap {unit_label}"

    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    header_font  = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    border_side  = Side(style="thin", color="CCCCCC")
    cell_border  = Border(left=border_side, right=border_side,
                          top=border_side, bottom=border_side)

    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value     = f"Rekap Absensi Murid {unit_label}"
    title_cell.font      = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment = center_align
    if start and end:
        ws.merge_cells("A2:H2")
        ws["A2"].value     = f"Periode: {start} s/d {end}"
        ws["A2"].alignment = center_align
        ws["A2"].font      = Font(italic=True, color="666666")

    headers = ["No", "Nama Murid", "Kelas", "Unit",
               "Hadir", "Izin", "Sakit", "Total"]
    start_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = cell_border

    status_colors = {"Hadir": "D4EDDA", "Izin": "FFF3CD", "Sakit": "F8D7DA"}
    for i, row in enumerate(rows, 1):
        r = start_row + i
        values = [i, excel_safe(row["nama"]), excel_safe(row["kelas"]),
                  excel_safe(row["unit"]),
                  row["hadir"], row["izin"], row["sakit"], row["total"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = center_align
            cell.border    = cell_border

    col_widths = [5, 25, 22, 8, 8, 8, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = f"A{start_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"rekap_{unit_label}_{start or 'awal'}_{end or 'akhir'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )

# ── Guru: materi ────────────────────────────────────────────────────────────
class MateriIn(BaseModel):
    judul: str
    isi: str

@app.get("/api/materi")
def list_materi(payload: dict = Depends(verify_token)):
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute("SELECT id, judul, isi, tanggal, created_by, created_at FROM materi ORDER BY tanggal DESC, id DESC")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

@app.post("/api/materi")
def create_materi(data: MateriIn, payload: dict = Depends(verify_token)):
    if not data.judul.strip() or not data.isi.strip():
        raise HTTPException(status_code=400, detail="Judul & isi materi wajib diisi")
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute(
        "INSERT INTO materi (judul, isi, tanggal, created_by) VALUES (?, ?, ?, ?)",
        (data.judul.strip(), data.isi.strip(), datetime.now().strftime("%Y-%m-%d"), payload["sub"]))
    conn.commit()
    mid = cur.lastrowid
    conn.close()
    return {"ok": True, "id": mid}

@app.delete("/api/materi/{materi_id}")
def delete_materi(materi_id: int, payload: dict = Depends(verify_token)):
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute(
        "DELETE FROM materi WHERE id = ? AND (created_by = ? OR ? = 'ALL')",
        (materi_id, payload["sub"], payload["unit"]))
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    if not deleted:
        raise HTTPException(status_code=404, detail="Materi tidak ditemukan")
    return {"ok": True}

# ── Guru: nilai ─────────────────────────────────────────────────────────────
class NilaiIn(BaseModel):
    murid_id: int
    mapel: str
    nilai: float
    tanggal: str = ""

@app.get("/api/nilai")
def list_nilai(murid_id: int = None, payload: dict = Depends(verify_token)):
    unit = payload["unit"]
    conn = get_conn()
    cur  = conn.cursor()
    if murid_id:
        cur.execute("""
            SELECT n.id, n.murid_id, n.mapel, n.nilai, n.tanggal
            FROM nilai n JOIN murid m ON m.id = n.murid_id
            WHERE n.murid_id = ? AND (m.unit = ? OR ? = 'ALL')
            ORDER BY n.tanggal DESC, n.id DESC
        """, (murid_id, unit, unit))
    else:
        cur.execute("""
            SELECT n.id, n.murid_id, n.mapel, n.nilai, n.tanggal
            FROM nilai n JOIN murid m ON m.id = n.murid_id
            WHERE m.unit = ? OR ? = 'ALL'
            ORDER BY n.tanggal DESC, n.id DESC
        """, (unit, unit))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

@app.post("/api/nilai")
def create_nilai(data: NilaiIn, payload: dict = Depends(verify_token)):
    if not data.mapel.strip():
        raise HTTPException(status_code=400, detail="Mata pelajaran wajib diisi")
    if not (0 <= data.nilai <= 100):
        raise HTTPException(status_code=400, detail="Nilai harus 0-100")
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute("SELECT id FROM murid WHERE id = ? AND (unit = ? OR ? = 'ALL')",
                (data.murid_id, payload["unit"], payload["unit"]))
    if not cur.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Murid tidak ditemukan")
    tanggal = data.tanggal or datetime.now().strftime("%Y-%m-%d")
    cur.execute(
        "INSERT INTO nilai (murid_id, mapel, nilai, tanggal) VALUES (?, ?, ?, ?)",
        (data.murid_id, data.mapel.strip(), data.nilai, tanggal))
    conn.commit()
    nid = cur.lastrowid
    conn.close()
    return {"ok": True, "id": nid}

@app.get("/api/nilai/rata2")
def rata2_nilai(payload: dict = Depends(verify_token)):
    unit = payload["unit"]
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute("""
        SELECT m.id AS murid_id, m.nama, m.kelas,
               ROUND(AVG(n.nilai), 2) AS rata2, COUNT(n.id) AS jumlah
        FROM murid m LEFT JOIN nilai n ON n.murid_id = m.id
        WHERE m.unit = ? OR ? = 'ALL'
        GROUP BY m.id ORDER BY m.urutan, m.nama
    """, (unit, unit))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

@app.delete("/api/nilai/{nilai_id}")
def delete_nilai(nilai_id: int, payload: dict = Depends(verify_token)):
    unit = payload["unit"]
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute("""
        DELETE FROM nilai WHERE id = ? AND murid_id IN (
            SELECT id FROM murid WHERE unit = ? OR ? = 'ALL'
        )
    """, (nilai_id, unit, unit))
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    if not deleted:
        raise HTTPException(status_code=404, detail="Nilai tidak ditemukan")
    return {"ok": True}
